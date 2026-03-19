"""Generación de reporte Excel de cuotas."""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from schoolai.db.models.cuota import Actividad, ActividadParticipante


def _hex_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def export_actividad_excel(
    actividad: Actividad,
    participantes: list[ActividadParticipante],
) -> bytes:
    """Genera un archivo Excel con el estado de pagos de una actividad.

    Retorna los bytes del archivo .xlsx.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Cuotas"

    # ── Encabezado ────────────────────────────────────────────────────────────
    monto_total = float(actividad.monto)
    ws.merge_cells("A1:F1")
    ws["A1"] = f"CUOTAS — {actividad.nombre.upper()}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Monto total: ${monto_total:.2f}   |   Exportado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Títulos de columnas ───────────────────────────────────────────────────
    headers = ["#", "Apellidos", "Nombres", "Total pagado", "Pendiente", "Estado"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _hex_fill("2D6A4F")
        cell.alignment = Alignment(horizontal="center")

    # ── Filas de estudiantes ──────────────────────────────────────────────────
    green = _hex_fill("D8F3DC")
    red   = _hex_fill("FFCDD2")
    yellow = _hex_fill("FFF9C4")

    for i, p in enumerate(participantes, 1):
        student   = p.student
        last_name  = getattr(student, "last_name",  "") or ""
        first_name = getattr(student, "first_name", "") or ""
        total      = float(p.total_pagado or 0)
        pendiente  = max(0.0, monto_total - total)

        if p.is_complete:
            estado = "✅ Completo"
            fill   = green
        elif total > 0:
            estado = f"⚠️ Parcial (${total:.2f})"
            fill   = yellow
        else:
            estado = "❌ Pendiente"
            fill   = red

        row = i + 4
        values = [i, last_name.title(), first_name.title(), f"${total:.2f}", f"${pendiente:.2f}", estado]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center" if col in (1, 4, 5, 6) else "left")

    # ── Resumen ───────────────────────────────────────────────────────────────
    total_row = len(participantes) + 6
    completos  = sum(1 for p in participantes if p.is_complete)
    parciales  = sum(1 for p in participantes if not p.is_complete and float(p.total_pagado or 0) > 0)
    pendientes = len(participantes) - completos - parciales

    ws.cell(row=total_row, column=1, value="RESUMEN").font = Font(bold=True)
    ws.cell(row=total_row + 1, column=1, value=f"✅ Completos: {completos}")
    ws.cell(row=total_row + 2, column=1, value=f"⚠️ Parciales: {parciales}")
    ws.cell(row=total_row + 3, column=1, value=f"❌ Pendientes: {pendientes}")
    ws.cell(row=total_row + 4, column=1, value=f"Total: {len(participantes)}")

    # ── Ancho de columnas ─────────────────────────────────────────────────────
    widths = [5, 20, 18, 14, 12, 18]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
